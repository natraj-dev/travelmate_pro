"""
Reports & Export — generates PDF / CSV / Excel files from tabular data.
"""
import csv
import io
import os
import uuid
from datetime import datetime

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.config import settings

REPORTS_DIR = os.path.join(settings.UPLOAD_DIR, "reports")


def _ensure_dir() -> None:
    os.makedirs(REPORTS_DIR, exist_ok=True)


def export_csv(title: str, headers: list[str], rows: list[list]) -> str:
    _ensure_dir()
    filename = f"{title.lower().replace(' ', '_')}_{uuid.uuid4().hex[:8]}.csv"
    path = os.path.join(REPORTS_DIR, filename)
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    return path


def export_excel(title: str, headers: list[str], rows: list[list]) -> str:
    _ensure_dir()
    filename = f"{title.lower().replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"
    path = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31] or "Report"

    ws.append(headers)
    header_fill = PatternFill(start_color="0F3330", end_color="0F3330", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for row in rows:
        ws.append(row)

    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else "A"].width = 22

    wb.save(path)
    return path


def export_pdf(title: str, headers: list[str], rows: list[list]) -> str:
    _ensure_dir()
    filename = f"{title.lower().replace(' ', '_')}_{uuid.uuid4().hex[:8]}.pdf"
    path = os.path.join(REPORTS_DIR, filename)

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(15, 51, 48)
    pdf.cell(0, 12, "TravelMate Pro", ln=True)
    pdf.set_font("Helvetica", "", 11)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 8, f"{title}  |  Generated {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", ln=True)
    pdf.ln(4)

    col_width = 277 / max(len(headers), 1)
    pdf.set_font("Helvetica", "B", 9)
    pdf.set_fill_color(15, 51, 48)
    pdf.set_text_color(255, 255, 255)
    for h in headers:
        pdf.cell(col_width, 8, str(h), border=1, fill=True)
    pdf.ln()

    pdf.set_font("Helvetica", "", 8)
    pdf.set_text_color(30, 36, 34)
    fill = False
    for row in rows:
        pdf.set_fill_color(246, 242, 233) if fill else pdf.set_fill_color(255, 255, 255)
        for value in row:
            pdf.cell(col_width, 7, str(value)[:40], border=1, fill=True)
        pdf.ln()
        fill = not fill

    pdf.output(path)
    return path
